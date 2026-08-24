"""Kuruma gönderilecek işaretli kopyayı üretir.

KAYNAK.md §2.10'da kuruma şu öneriliyor: Düzeltmeler uygulanmış bir kopya,
**yalnız değişen hücreler işaretli**, aslına dokunulmadan. Bu betik o kopyayı
üretiyor.

Neden ayrı bir betik, test değil: Kaynak tablo `lokal/` altında ve depoya
girmiyor (CONTRIBUTING.md, 2. kural). Yani bu iş CI'da koşamaz. Testler çevrimdışı ve
kurulumsuz kalıyor, bu betik elle çalıştırılıyor.

    python3 scripts/kuruma_geri_bildirim.py maku

Çıktı `lokal/<uni>/geri-bildirim-<tarih>.xlsx` olarak yazılıyor · o da depoya
girmiyor.

NE YAPMIYOR

Kaynağın kendisine dokunmuyor. Okuyup yeni bir dosya yazıyor, aslı olduğu yerde
kalıyor. Bu, "kaynak veri asla yerinde değiştirilmez" ilkesinin uygulaması ·
her düzeltme kaynakta değil, okurken yapılıyor ve iz bırakıyor.
"""
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "scripts"))

# Fill colour for changed cells. This is a notice, not a warning: a calm
# yellow that stays readable on screen and still prints distinguishably.
ISARET = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
BASLIK = Font(bold=True)


def yukle_kayit(uni_id):
    import build_data as bd
    for u in bd.UNIVERSITIES:
        if u["id"] == uni_id:
            return bd, u
    raise SystemExit(f"'{uni_id}' diye bir üniversite kaydı yok. "
                     f"Tanımlı olanlar: {[u['id'] for u in bd.UNIVERSITIES]}")


def uret(uni_id):
    import json
    bd, uni = yukle_kayit(uni_id)

    veri_yolu = ROOT / "site" / f"data-{uni_id}.json"
    if not veri_yolu.exists():
        raise SystemExit(f"{veri_yolu.name} yok. Önce build_data.py çalıştırılmalı.")
    kayitlar = json.loads(veri_yolu.read_text(encoding="utf-8"))["agreements"]

    kaynak = uni["local"]
    if not kaynak.exists():
        raise SystemExit(f"Kaynak tablo bulunamadı: {kaynak}")

    wb = openpyxl.load_workbook(kaynak)
    ws = wb.worksheets[uni.get("sheet", 0)]
    # The column map is university-specific. It comes from the same function
    # the parser uses, so the two cannot drift apart.
    harita_islevi = getattr(bd, f"{uni_id}_sutunlari", None)
    if harita_islevi is None:
        raise SystemExit(f"'{uni_id}' için sütun haritası işlevi yok "
                         f"(beklenen ad: {uni_id}_sutunlari).")
    S = harita_islevi(ws)

    # Records must be matched back to their source rows. The shared key is
    # university name + Erasmus code. That pair can repeat in the sheet, so we
    # take the first matching row that has NOT been used yet.
    kullanilan = set()
    isaretlenen = 0
    aciklama = {
        "erasmus-kodu": "Ülke, Erasmus kurum kodunun ön ekinden alındı.",
        "iptal-notu": "İptal edildiği yazılı öğrenim kademesi listeden çıkarıldı.",
        "eksik-sifir": "Üç haneli alan kodu, baştaki sıfır düşmüş varsayılarak tamamlandı.",
        "isced-etiketi": "Türkçe bölüm adı boş olduğu için İngilizce ISCED etiketi gösteriliyor.",
        "yazim-birligi": "Aynı bölümün birden çok yazımı var; en sık kullanılanı gösteriliyor.",
    }
    # Which kind of trace lands in which column. The distinction is subtle:
    # `iscedFamily` concerns the field CODE, `department` the field NAME. They
    # sit in different columns, and swapping them would point the institution
    # at the wrong cell.
    SUTUN = {"country": "ulke", "levels": "eqf",
             "iscedFamily": "isced_kodu", "department": "bolum_tr"}

    for rec in kayitlar:
        iz = rec.get("sourceDiff")
        if not iz:
            continue
        anahtar = (rec.get("university", "").strip().lower(),
                   rec.get("erasmusCode", "").strip().lower())
        for r in range(2, ws.max_row + 1):
            if r in kullanilan:
                continue
            u = str(ws.cell(r, S["universite"]).value or "").strip().lower()
            k = str(ws.cell(r, S["erasmus_kodu"]).value or "").strip().lower()
            if (u, k) != anahtar:
                continue
            kullanilan.add(r)
            for tur, ayrinti in iz.items():
                sutun = S.get(SUTUN.get(tur, ""))
                if not sutun:
                    continue
                hucre = ws.cell(r, sutun)
                hucre.fill = ISARET
                kaynakta = ayrinti.get("kaynakta")
                metin = aciklama.get(ayrinti.get("sebep"), "Kaynaktan ayrıldı.")
                if kaynakta:
                    metin += f"\nKaynakta: {kaynakta}"
                hucre.comment = Comment(metin, "Exchange Atlas")
                isaretlenen += 1
            break

    # Cover sheet, so whoever opens the file sees what it is right away
    bilgi = wb.create_sheet("Exchange Atlas · açıklama", 0)
    satirlar = [
        ("Bu dosya nedir",),
        ("Yayımladığınız listenin bir kopyası. Yalnız RENKLİ hücreler bizim "
         "okuma kararımızla farklı gösteriliyor.",),
        ("Renksiz hücrelere dokunulmadı.",),
        ("",),
        ("Ne yapmanız gerekiyor",),
        ("Tabloyu baştan sona okumanız gerekmiyor. Yalnız renkli hücrelere "
         "bakın; her birinin üzerinde sebebi yazılı bir not var.",),
        ("Kabul ya da reddetmek sizde. Biz hangi yazımın doğru olduğuna karar "
         "vermiyoruz.",),
        ("",),
        ("Aslınız değişmedi",),
        ("Bu ayrı bir dosya. Kendi listeniz elinizde olduğu gibi duruyor.",),
        ("",),
        (f"Üretildiği tarih: {datetime.now().date().isoformat()}",),
        (f"İşaretlenen hücre: {isaretlenen}",),
    ]
    for i, s in enumerate(satirlar, start=1):
        h = bilgi.cell(i, 1, s[0])
        if s[0] and not s[0].startswith(("Bu ", "Tabloyu", "Kabul", "Renksiz",
                                         "Yayımladığınız", "Üretildiği", "İşaretlenen")):
            h.font = BASLIK
    bilgi.column_dimensions["A"].width = 96

    cikti = kaynak.parent / f"geri-bildirim-{datetime.now().date().isoformat()}.xlsx"
    wb.save(cikti)
    print(f"{uni_id}: {isaretlenen} hücre işaretlendi -> {cikti}", file=sys.stderr)
    return cikti


if __name__ == "__main__":
    if len(sys.argv) != 2:
        raise SystemExit("Kullanım: python3 scripts/kuruma_geri_bildirim.py <uni-id>")
    uret(sys.argv[1])
