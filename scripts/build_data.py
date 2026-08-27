#!/usr/bin/env python3
"""
İkili anlaşma listeleri -> site/data-<uni>.json + site/universities.json

Çok üniversiteli: her üniversitenin kendi Excel formatı için ayrı parser var,
hepsi aynı kayıt şemasına normalize edilir. Yeni üniversite = UNIVERSITIES'e kayıt +
parser fonksiyonu.

  Yerel snapshot'lardan (varsayılan, demo için güvenli):
      python3 scripts/build_data.py
  Canlı kaynaklardan çek:
      python3 scripts/build_data.py --pull
"""
import json
import re
import sys
import unicodedata
import urllib.request
from datetime import datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
SITE = ROOT / "site"

# ISCED-F 2013 broad fields (first two digits of the code)
ISCED_FAMILY = {
    "00": ("Genel programlar", "Generic programmes"),
    "01": ("Eğitim", "Education"),
    "02": ("Sanat ve beşeri bilimler", "Arts and humanities"),
    "03": ("Sosyal bilimler, gazetecilik", "Social sciences, journalism"),
    "04": ("İşletme, yönetim, hukuk", "Business, administration, law"),
    "05": ("Doğa bilimleri, matematik", "Natural sciences, mathematics"),
    "06": ("Bilişim ve iletişim (BİT)", "Information and Communication Tech."),
    "07": ("Mühendislik, üretim, inşaat", "Engineering, manufacturing, construction"),
    "08": ("Tarım, ormancılık, veterinerlik", "Agriculture, forestry, veterinary"),
    "09": ("Sağlık ve refah", "Health and welfare"),
    "10": ("Hizmetler", "Services"),
}

# Unicode decomposition handles characters that split into "letter + mark"
# (ä -> a + ¨). The ones below are separate letters, so that route does not work;
# they are mapped by hand. Must stay identical to its twin in app.js
# (see tests/test_couplings.py).
FOLD_EXCEPTIONS = str.maketrans({
    "ı": "i", "ł": "l", "Ł": "L", "đ": "d", "Đ": "D",
    "ø": "o", "Ø": "O", "ß": "ss",
})


def fold(s):
    """Arama anahtarı: aksan ve işaret farklarını düzleştirir."""
    if s is None:
        return ""
    s = str(s).translate(FOLD_EXCEPTIONS)
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return s.lower()


def tr_upper(s):
    """Türkçe-doğru büyük harf: 'Bulgaristan' -> 'BULGARİSTAN'."""
    return s.replace("i", "İ").replace("ı", "I").upper()


def unify_department_spelling(agreements):
    """Aynı bölümün yalnız büyük/küçük harfle ayrışan yazımlarını birleştirir.

    Kaynak tabloda aynı bölüm birden çok yazımla giriliyor: `İnşaat
    Mühendisliği` ve `İnşaat mühendisliği` gibi. Arama bunu zaten
    düzleştiriyor (`fold` büyük/küçük farkını siliyor), etkilenen tek şey
    **kartta görünen yazım**: Aynı bölüm iki farklı biçimde görünüyor.

    Seçim kuralı: **kurumun kendi yazımlarından en sık kullanılanı.** Yeni bir
    yazım üretilmiyor, hangisinin "doğru" olduğuna karar verilmiyor; yalnız
    kurumun ağırlıklı olarak kullandığı biçim tutarlı gösteriliyor.

    **Beraberlikte dokunulmuyor.** İki yazım eşit sıklıktaysa seçmek için
    elimizde ölçü kalmıyor ve orada tercih yapmak yazım tercihi dayatmak
    olurdu. O gruplar olduğu gibi bırakılıyor ve `KAYNAK.md`'de bildiriliyor.

    Her değişiklik `sourceDiff`'e yazılıyor, yani öğrenci kaynakta ne yazdığını
    kartta görüyor.
    """
    def tr_lower(s):
        return s.replace("I", "ı").replace("İ", "i").lower()

    gruplar = {}
    for a in agreements:
        dep = a.get("department")
        # If the department name already carries a trace (it fell back to the
        # ISCED label), that spelling is not the institution's own; keep it out.
        if not dep or "department" in (a.get("sourceDiff") or {}):
            continue
        gruplar.setdefault(tr_lower(dep), {})
        gruplar[tr_lower(dep)][dep] = gruplar[tr_lower(dep)].get(dep, 0) + 1

    secilen, beraberlik = {}, 0
    for anahtar, sayim in gruplar.items():
        if len(sayim) < 2:
            continue
        sirali = sorted(sayim.items(), key=lambda t: -t[1])
        if sirali[0][1] == sirali[1][1]:
            beraberlik += 1
            continue
        secilen[anahtar] = sirali[0][0]

    degisen = 0
    for a in agreements:
        dep = a.get("department")
        if not dep or "department" in (a.get("sourceDiff") or {}):
            continue
        kazanan = secilen.get(tr_lower(dep))
        if not kazanan or kazanan == dep:
            continue
        a.setdefault("sourceDiff", {})["department"] = {
            "kaynakta": dep, "sebep": "yazim-birligi",
        }
        a["department"] = kazanan
        degisen += 1
    return degisen, beraberlik


def clean(v):
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.isoformat()
    # Non-breaking space (\xa0) becomes a normal space: indistinguishable by eye
    # but it breaks search matching. .strip() already removes it at the START/END;
    # this conversion is for the ones left in the MIDDLE of the text.
    return str(v).replace("\xa0", " ").strip()


def isced_family(code):
    if code is None:
        return None
    digits = re.sub(r"\D", "", str(code))
    if not digits:
        return None
    if len(digits) == 1:
        digits = "0" + digits
    # Broad field = first two digits. Excel can eat a leading zero
    # ("0610" -> "610"), so if the first two digits are NOT a valid family we
    # assume a lost zero and prepend "0". If they ARE valid ("104" -> "10",
    # Services) LEAVE IT ALONE: otherwise genuine three-digit codes in family 10
    # would silently shift into "01".
    fam = digits[:2]
    if fam not in ISCED_FAMILY:
        fam = ("0" + digits)[:2]
    return fam if fam in ISCED_FAMILY else None


def aile_tahminle_bulundu(code):
    """Aile, baştaki sıfırın yendiği VARSAYILARAK mı bulundu?

    Üç haneli bir kodda iki okuma mümkün: `104` ya "10" ailesi (Hizmetler) ya da
    baştaki sıfırı yenmiş bir `0104`. Kod ilk okumayı deniyor, tutmazsa sıfır
    ekliyor. İkinci yola düşen kayıtlar **tahminle** sınıflandırılmış oluyor ve
    bunu öğrencinin görmesi gerekiyor.

    Bugün 12 kayıt bu durumda ve on ikisi de bölüm adına bakılarak elle
    doğrulandı (bkz. tests/test_kaynak_verisi.py, UcHaneliKodlar).
    """
    if code is None:
        return False
    digits = re.sub(r"\D", "", str(code))
    if len(digits) != 3:
        return False
    return digits[:2] not in ISCED_FAMILY


def fam_fields(code):
    fam = isced_family(code)
    return {
        "iscedFamily": fam,
        "iscedFamilyTr": ISCED_FAMILY[fam][0] if fam else None,
        "iscedFamilyEn": ISCED_FAMILY[fam][1] if fam else None,
    }


def parse_validity(text):
    raw = clean(text)
    return {"raw": raw} if raw else None


# The prefix of an Erasmus institution code identifies the country. Two shapes
# exist: today's two-letter form (ISO 3166-1, with two EU exceptions: Greece EL,
# United Kingdom UK) and the pre-2014 single-letter form.
#
# NOTE ON EVIDENCE: the two-letter convention is well known, but it was not
# verified here against an official specification. What IS verified is the
# measurement below, taken from the data itself. KAYNAK.md says the same;
# the two must not drift apart.
#
# Measured across the two sources that HAVE a country column (MAKU 468 rows,
# ESOGU 259 rows): 722 rows agree, 4 disagree, 2 have an empty column. The 4
# disagreements are all MAKU's and all documented; ESOGU agrees on every row.
#
# The third source (Marmara) has no country column at all, so there the prefix
# is not a cross-check but the ONLY reading. That is why a Marmara row with an
# unreadable prefix is dropped rather than shown with a blank country.
ERASMUS_ULKE = {
    "AT": "Avusturya", "BE": "Belçika", "BG": "Bulgaristan", "CY": "Kıbrıs",
    "CZ": "Çekya", "DE": "Almanya", "DK": "Danimarka", "EE": "Estonya",
    "EL": "Yunanistan", "ES": "İspanya", "FI": "Finlandiya", "FR": "Fransa",
    "HR": "Hırvatistan", "HU": "Macaristan", "IE": "İrlanda", "IS": "İzlanda",
    "IT": "İtalya", "LT": "Litvanya", "LV": "Letonya", "MK": "Kuzey Makedonya",
    "MT": "Malta", "NL": "Hollanda", "NO": "Norveç", "PL": "Polonya",
    "PT": "Portekiz", "RO": "Romanya", "RS": "Sırbistan", "SE": "İsveç",
    "SI": "Slovenya", "SK": "Slovakya", "TR": "Türkiye", "UK": "Birleşik Krallık",
    # Pre-2014 single-letter codes
    "A": "Avusturya", "B": "Belçika", "D": "Almanya", "E": "İspanya",
    "F": "Fransa", "G": "Yunanistan", "I": "İtalya", "P": "Portekiz",
    # S (Sverige) was missing until Marmara's list arrived: five rows were
    # dropped for an unreadable country. Three witnesses agree it is Sweden ·
    # the legacy pattern (a letter from the country's own name, like SF for
    # Suomi Finland), and the institution names themselves: Kristianstad,
    # Södertörn, Stockholm, Umeå.
    "S": "İsveç",
    "SF": "Finlandiya",
}


def ulke_onekten(erasmus_kodu):
    """Erasmus kurum kodundan ülkeyi çıkarır; bilinmiyorsa None."""
    m = re.match(r"\s*([A-Z]{1,2})[\s-]", erasmus_kodu or "")
    return ERASMUS_ULKE.get(m.group(1)) if m else None


def isced_kodu(*hucreler):
    """Aynı kodu birden çok sütun taşıyor; en ayrıntılı olanı seçer.

    MAKÜ tablosunda ISCED kodu üç yerde geçiyor: bölüm adlarının önünde
    ("0610-Bilgisayar Mühendisliği") ve ayrı bir kod sütununda. Uzun süre
    yalnız kod sütunu okundu ve o sütun 467 kaydın 246'sında boştu; site de
    o kayıtlar için "alan bilgisi yok" gösteriyordu.

    Ölçüldü: boş görünen 246 kaydın 245'inde kod bölüm adının önünde duruyordu.
    Yani eksik olan veri değil, okuduğumuz sütundu.

    ISCED hiyerarşik: 06 geniş alan, 061 dar alan, 0610 ayrıntılı alan. Üç
    sütun farklı derinlik taşıyabiliyor ve bu bir çelişki değil. En uzun olanı
    seçiyoruz, çünkü ayrıntılı olan dar olanı zaten içeriyor.

    Sıra da önemli: eşit uzunlukta iki aday varsa Türkçe sütun kazanıyor,
    çünkü sitede gösterilen bölüm adı oradan geliyor.
    """
    candidates = []
    for h in hucreler:
        m = re.match(r"\s*(\d{3,4})\b", clean(h))
        if m:
            candidates.append(m.group(1))
    return max(candidates, key=len) if candidates else ""


# Free-text cells sometimes carry an e-mail address. The published data must
# not (KAYNAK.md §5): we do not republish addresses that people did not give
# us, and an address in a partner's note is no different from one in a signer
# column, which is not read at all.
#
# The note itself is kept, because it carries real information ("B1 German
# required"). Only the address is removed, and the removal leaves a trace, so
# a student does not read a truncated sentence and wonder what is missing.
EPOSTA = re.compile(r"[\w.+-]+@[\w.-]+\.[A-Za-z]{2,}")


def eposta_ayikla(metin):
    """Returns (cleaned text, was anything removed)."""
    if not metin or not EPOSTA.search(metin):
        return metin, False
    temiz = re.sub(r"\s{2,}", " ", EPOSTA.sub("", metin)).strip(" ·,;")
    return temiz, True


def make_search(rec, extra=""):
    return fold(" ".join(filter(None, [
        rec["country"], rec["university"], rec["department"],
        rec["iscedCode"], rec["erasmusCode"],
        rec["iscedFamilyTr"] or "", rec["iscedFamilyEn"] or "", extra,
    ])))


# ═══════════════════════════════ MAKU ════════════════════════════════

# Columns are located by NAME, not by index. The reason was learned by
# measurement: when a column is inserted into the source sheet, index-based code
# does not crash, it **produces meaningless data** -- it finds a department name
# in the country column and upper-cases it.
#
# This debt was paid twice. The ISCED code was read from the wrong column (243
# records ended up with no field) and so was the country (4 records appeared in
# the wrong country, 2 did not appear at all). In both cases reading by index
# made the question "what is in which column" hard to even ask.
MAKU_BASLIKLAR = {
    "universite":   "İkili Anlaşma Yapılan Yüksek Öğretim Kurumu",
    "gecerlilik":   "Geçerlilik Süresi",
    "erasmus_kodu": "Üniversite ID Kodu",
    "aciklama":     "Açıklama",
    "isced_kodu":   "Ders Alan Kodu",
    "dil":          "Eğitim Dili",
    "ulke":         "Ülke",
    "kont_ogrenim": "Öğrenim Hareketliliği Kontenjanı",
    "kont_ders":    "Ders Verme Kontenjanı",
    "kont_staj":    "Staj Hareketliliği Kontenjanı",
    "kont_egitim":  "Eğitim Alma Kontenjanı",
}

# The EQF header is long and carries an explanation inside it
# ("... EQF 5: ... EQF 6: ..."), so exact matching would be brittle; a prefix
# match is enough.
MAKU_BASLIK_ONEK = {"eqf": "Öğrenim Kademesi"}


def _bosluk_sadelestir(s):
    return re.sub(r"\s+", " ", str(s or "")).strip()


def maku_sutunlari(ws):
    """Başlık satırından ada göre sütun numaralarını bulur.

    İki sütun **birebir aynı başlığı** taşıyor: F ve H, ikisi de "İkili Anlaşma
    Yapılan Bölüm Ders Alan Kodu (ISCED)". Bu, borcun neden uzun süre
    ödenmediğinin de sebebi; ad tek başına yetmiyor.

    İkisi içerikleriyle ayrılıyor: F resmî İngilizce ISCED etiketini taşıyor
    (Türkçe harf oranı %1), H ise bölümün Türkçe adını (%72). Oran ölçüldü ve
    aradaki fark tesadüfe yer bırakmayacak kadar büyük.
    """
    baslik = {}
    ikili_isced = []
    for c in range(1, ws.max_column + 1):
        ad = _bosluk_sadelestir(ws.cell(1, c).value)
        if not ad:
            continue
        if "Bölüm Ders Alan Kodu" in ad:
            ikili_isced.append(c)
        else:
            baslik.setdefault(ad, c)

    sutun = {}
    for anahtar, ad in MAKU_BASLIKLAR.items():
        if ad not in baslik:
            raise SystemExit(f"maku: '{ad}' başlıklı sütun bulunamadı. Kaynak "
                             f"tablonun başlıkları değişmiş olabilir; "
                             f"scripts/build_data.py içindeki MAKU_BASLIKLAR "
                             f"gözden geçirilmeli.")
        sutun[anahtar] = baslik[ad]
    for anahtar, onek in MAKU_BASLIK_ONEK.items():
        eslesen = [c for a, c in baslik.items() if a.startswith(onek)]
        if not eslesen:
            raise SystemExit(f"maku: '{onek}...' ile başlayan sütun yok.")
        sutun[anahtar] = eslesen[0]

    if len(ikili_isced) != 2:
        raise SystemExit(f"maku: ISCED başlıklı sütun sayısı {len(ikili_isced)}, "
                         f"beklenen 2. Tablo düzeni değişmiş.")

    # Which one is the Turkish department name? Decided by the ratio of
    # Turkish-specific letters.
    TR_HARF = set("çğıöşüÇĞİÖŞÜ")
    oran = []
    for c in ikili_isced:
        trli = toplam = 0
        for r in range(2, ws.max_row + 1):
            ad = re.sub(r"^\s*\d{3,4}[\s.\-]*", "", clean(ws.cell(r, c).value))
            if not ad:
                continue
            toplam += 1
            if TR_HARF & set(ad):
                trli += 1
        oran.append(trli / toplam if toplam else 0)
    if abs(oran[0] - oran[1]) < 0.2:
        raise SystemExit(f"maku: İki ISCED sütunu içerikle ayrılamadı "
                         f"(Türkçe harf oranları {oran[0]:.2f} ve {oran[1]:.2f}). "
                         f"Ayrım belirsiz; elle bakılmalı.")
    tr_index = 0 if oran[0] > oran[1] else 1
    sutun["bolum_tr"] = ikili_isced[tr_index]
    sutun["isced_etiket_en"] = ikili_isced[1 - tr_index]
    return sutun


def parse_maku(ws):
    """KA131 'Güncel Anlaşmalar': 19 sütun, satır 2'den itibaren.

    Sütunlar: C üni, D geçerlilik, E kod, F ISCED etiketi (EN), G açıklama,
    H bölüm adı (TR), I ISCED kodu, J eğitim dili, K EQF seviyeleri, L ülke,
    M öğrenim, N ders verme, O staj, P eğitim alma.

    Q ve R sütunları
    (imzalayan kişi ve e-postası) BİLEREK okunmuyor · bkz. `KAYNAK.md §5`.

    F ile H'nin başında da ISCED kodu duruyor ve I çoğu satırda boş; kod
    üçünden birlikte okunuyor (bkz. isced_kodu)."""
    S = maku_sutunlari(ws)
    agreements = []
    skipped = 0
    struck_out = 0
    ulke_duzeltme, ulke_tamamlama = [], []
    for r in range(2, ws.max_row + 1):
        def col(ad):
            return ws.cell(row=r, column=S[ad]).value
        university = clean(col("universite"))
        # A name with no letters is a data error (the source has a cell holding '0')
        if not university or not re.search(r"[A-Za-zĞÜŞİÖÇğüşıöç]", university):
            continue

        # A struck-through row is an agreement deleted at the source. The row was
        # never removed from the sheet, only marked by formatting, so a reader
        # looking at cell values cannot see it. This reached production once: an
        # agreement no longer valid sat in the list and a student could apply.
        if any((ws.cell(row=r, column=c).font or {}) and ws.cell(row=r, column=c).font.strike
               for c in (S["isced_etiket_en"], S["bolum_tr"])):
            struck_out += 1
            continue

        # Country can be read from two places: the country column and the prefix
        # of the Erasmus code. The prefix follows a fixed convention while the
        # column is typed by hand, so when they disagree the prefix wins and the
        # correction is recorded.
        country = clean(col("ulke"))
        onekten = ulke_onekten(clean(col("erasmus_kodu")))
        # Every departure from the source is recorded; the student sees it on the card
        iz = {}
        if onekten and country and tr_upper(onekten) != tr_upper(country):
            ulke_duzeltme.append((clean(col("universite"))[:44], country, onekten))
            iz["country"] = {"kaynakta": country, "sebep": "erasmus-kodu"}
            country = onekten
        elif not country:
            if not onekten:
                skipped += 1
                continue
            ulke_tamamlama.append((clean(col("universite"))[:44], onekten))
            iz["country"] = {"kaynakta": None, "sebep": "erasmus-kodu"}
            country = onekten
        country = tr_upper(country)

        # EQF -> level flags (look at digits only, to survive typos like 'QQF 6')
        #
        # Cancelled levels are stripped first: the source has a cell reading
        # "EQF 6 / EQF 7 (EQF 7 CANCELLED)". A plain digit scan would also find 7
        # and make the master's level look open. Any level marked cancelled inside
        # parentheses is dropped.
        eqf_raw = clean(col("eqf"))
        cancelled = set()
        for chunk in re.findall(r"\(([^)]*)\)", eqf_raw):
            # The source sheet is in Turkish; the searched words are its own text.
            if re.search(r"iptal|geçersiz|kaldır", chunk, re.I):
                cancelled.update(re.findall(r"[5678]", chunk))
        eqf = set(re.findall(r"[5678]", re.sub(r"\([^)]*\)", "", eqf_raw))) - cancelled
        if cancelled:
            iz["levels"] = {"kaynakta": eqf_raw, "sebep": "iptal-notu"}
        levels = {}
        for d, key in (("5", "onlisans"), ("6", "lisans"),
                       ("7", "yukseklisans"), ("8", "doktora")):
            if d in eqf:
                levels[key] = True

        # Strip the ISCED code prefix from names ("0610-Computer Eng." ->
        # "Computer Eng."); the code is read separately and shown as a chip.
        #
        # The two columns do not carry the same thing: the Turkish column holds
        # the real department name, while the English column holds the official
        # ISCED
        # etiketi ("Training for pre-school teachers", "Veterinary", "Education").
        # Measured: across 468 filled cells there are only 117 distinct values and
        # the most frequent ones are ISCED labels. Hence the name is not "dept_en".
        dept_tr = re.sub(r"^\s*\d{3,4}\s*-\s*", "", clean(col("bolum_tr")).strip())
        isced_label_en = re.sub(r"^\s*\d{3,4}\s*-\s*", "", clean(col("isced_etiket_en")).strip())
        lang_raw = re.sub(r"\s*\n\s*", " ", clean(col("dil"))).strip()
        not_temiz, eposta_vardi = eposta_ayikla(clean(col("aciklama")))
        if eposta_vardi:
            iz["quotaNote"] = {"kaynakta": None, "sebep": "eposta-ayiklandi"}

        # When the Turkish department name is empty we fall back to the English
        # ISCED label. This fallback used to be SILENT: what looked like a
        # department name on the card was actually a field label, and the reader
        # could not tell. It now leaves a trace.
        if not dept_tr and isced_label_en:
            iz["department"] = {"kaynakta": None, "sebep": "isced-etiketi"}

        # Only FILLED fields are written. A missing field means "no data"; writing
        # a fixed empty value would add useless bytes to every record. The site
        # already reads a missing field correctly (a.website && ...).
        rec = {
            "country": country,
            "erasmusCode": clean(col("erasmus_kodu")),
            "validity": parse_validity(col("gecerlilik")),
            "university": university,
            "iscedCode": isced_kodu(col("bolum_tr"), col("isced_etiket_en"), col("isced_kodu")),
            "department": dept_tr or isced_label_en,
            "levels": levels,
            "quotaStudy": clean(col("kont_ogrenim")),
            "quotaInternship": clean(col("kont_staj")),
            "quotaStaffTeach": clean(col("kont_ders")),
            "quotaStaffTrain": clean(col("kont_egitim")),
            "quotaNote": not_temiz,
            # The 'language of instruction' column: not a requirement, shown raw
            "language": {"raw": lang_raw} if lang_raw else None,
            # Fields where we departed from the source. When empty the field is
            # omitted entirely: 461 of 468 records have no departure, so it should
            # not take up space.
            **({"sourceDiff": iz} if iz else {}),
        }
        rec.update(fam_fields(rec["iscedCode"]))
        # If the family was inferred, leave a trace: the student should know that
        # the field it appears under rests on an assumption.
        if rec["iscedFamily"] and aile_tahminle_bulundu(rec["iscedCode"]):
            rec.setdefault("sourceDiff", {})["iscedFamily"] = {
                "kaynakta": rec["iscedCode"], "sebep": "eksik-sifir",
            }
        rec["search"] = make_search(rec, extra=isced_label_en)
        agreements.append(rec)
    # Country corrections are recorded: fixing them silently would hide an error
    # that ought to be reported back to the source.
    for ad, yazan, dogru in ulke_duzeltme:
        print(f"  maku: ülke düzeltildi · {ad}: {yazan} -> {dogru} "
              f"(Erasmus kodunun ön eki esas alındı)", file=sys.stderr)
    for ad, dogru in ulke_tamamlama:
        print(f"  maku: ülke boştu, ön ekten dolduruldu · {ad}: {dogru}",
              file=sys.stderr)
    if skipped:
        print(f"  maku: ülkesi hiçbir yerden okunamayan {skipped} satır atlandı",
              file=sys.stderr)
    if struck_out:
        print(f"  maku: üstü çizili {struck_out} satır atlandı (kaynakta silinmiş)",
              file=sys.stderr)
    return agreements


# ══════════════════════════════ MARMARA ══════════════════════════════

# The sheet holds the quota AND the levels in one cell. Grammar, taken from
# the data (68 distinct spellings, all of them parse):
#
#   "4L"          4 places, bachelor's only
#   "2L,1YL"      2 bachelor's + 1 master's, counted separately
#   "2(L,YL,D)"   2 places SHARED across bachelor's, master's and doctorate
#   "0"           no student quota
#
# Commas inside parentheses bind the group, commas outside separate segments.
# A first attempt split on every comma and broke "2(L,YL,D)" into three
# fragments; that is why the split below is depth-aware.
MARMARA_SEVIYE = {"ÖL": "onlisans", "L": "lisans", "YL": "yukseklisans", "D": "doktora"}
MARMARA_PARCA = re.compile(r"^(\d+)\s*(?:\(([^)]*)\)|([A-ZÖ]*))$")


def marmara_kontenjan_bol(s):
    """Splits on commas OUTSIDE parentheses."""
    parcalar, derinlik, son = [], 0, 0
    for i, ch in enumerate(s):
        if ch == "(":
            derinlik += 1
        elif ch == ")":
            derinlik -= 1
        elif ch == "," and derinlik == 0:
            parcalar.append(s[son:i])
            son = i + 1
    parcalar.append(s[son:])
    return [p.strip() for p in parcalar if p.strip()]


def marmara_seviyeler(kontenjan):
    """Reads the level flags out of the quota cell.

    A level inside parentheses gets "shared" instead of a number, because the
    places are not per-level: "2(L,YL)" means two places that bachelor's and
    master's students compete for together. The site already renders that with
    a star, so the distinction reaches the student instead of being flattened
    into "2 for each".
    """
    seviyeler = {}
    for parca in marmara_kontenjan_bol(clean(kontenjan)):
        m = MARMARA_PARCA.match(parca)
        if not m:
            continue
        sayi, grup, tek = m.group(1), m.group(2), m.group(3)
        if int(sayi) <= 0:
            continue
        if grup:
            for kod in re.split(r"[,\s]+", grup):
                if kod in MARMARA_SEVIYE:
                    seviyeler[MARMARA_SEVIYE[kod]] = "shared"
        elif tek in MARMARA_SEVIYE:
            seviyeler[MARMARA_SEVIYE[tek]] = sayi
    return seviyeler


MARMARA_LEJANT_BASI = "Zorunlu Dil ID"


def marmara_dil_sozlugu(wb):
    """The language legend sits BELOW the data in every sheet.

    The 'foreign language' column holds numbers ("1", "1;19"), not names. The
    key is a two-column table appended under each sheet's agreements. Without
    it the field is unreadable, so it is not optional.

    Read from every sheet and cross-checked: 66 entries, no sheet disagrees
    with another. Had they disagreed we would have to say which one wins;
    they do not, so the question does not arise.
    """
    sozluk, catisma = {}, []
    for ws in wb.worksheets:
        lejantta = False
        for satir in ws.iter_rows(min_row=2, values_only=True):
            if not satir or satir[0] is None:
                continue
            if clean(satir[0]) == MARMARA_LEJANT_BASI:
                lejantta = True
                continue
            if not lejantta:
                continue
            kod, ad = clean(satir[0]), clean(satir[1]) if len(satir) > 1 else ""
            if kod and ad:
                if kod in sozluk and sozluk[kod] != ad:
                    catisma.append((ws.title, kod, sozluk[kod], ad))
                sozluk[kod] = ad
    if catisma:
        print(f"  marmara: dil sözlüğü sayfalar arasında ÇELİŞİYOR: {catisma[:3]}",
              file=sys.stderr)
    return sozluk


def parse_marmara(wb):
    """KA131 list: one sheet per faculty, 27 sheets, plus a language legend.

    Columns: A code, B university, C start, D end, E department,
    F study quota, G traineeship quota, H staff teaching, I staff training,
    J language (numeric), K note, L agreement type.

    TWO THINGS THIS SOURCE DOES NOT HAVE, and both are left empty rather
    than guessed:

      · No ISCED code anywhere · not in a column, not in front of the
        department name. So the field filter does not cover these records and
        the site shows them under "no field information". Inferring a code
        from the department name would be a guess presented as data.
      · No country column · the country comes only from the Erasmus code
        prefix. For MAKU the prefix is a cross-check against a column; here it
        is the single source, so a row whose prefix is unreadable is dropped
        rather than shown with a blank country.
    """
    diller = marmara_dil_sozlugu(wb)
    agreements = []
    atlanan = 0
    for ws in wb.worksheets:
        fakulte = ws.title.strip()
        for satir in ws.iter_rows(min_row=2, values_only=True):
            if not satir or satir[0] is None:
                continue
            if clean(satir[0]) == MARMARA_LEJANT_BASI:
                break  # the rest of this sheet is the legend, not agreements
            kod = clean(satir[0])
            university = clean(satir[1])
            if not university or not re.search(r"[A-Za-zĞÜŞİÖÇğüşıöç]", university):
                continue
            country = ulke_onekten(kod)
            if not country:
                atlanan += 1
                continue
            bas, bit = satir[2], satir[3]
            yil = [str(d.year) for d in (bas, bit) if hasattr(d, "year")]
            gecerlilik = "/".join(yil) if len(yil) == 2 else ""

            not_temiz, eposta_vardi = eposta_ayikla(clean(satir[10]))
            dil_ham = clean(satir[9])
            kodlar = [k.strip() for k in re.split(r"[;,]", dil_ham)]
            adlar = [diller[k] for k in kodlar if k in diller]
            rec = {
                "country": tr_upper(country),
                "erasmusCode": kod,
                "validity": {"raw": gecerlilik} if gecerlilik else None,
                "university": university,
                "iscedCode": "",
                "department": clean(satir[4]),
                "levels": marmara_seviyeler(satir[5]),
                "quotaStudy": clean(satir[5]),
                "quotaInternship": clean(satir[6]),
                "quotaStaffTeach": clean(satir[7]),
                "quotaStaffTrain": clean(satir[8]),
                "quotaNote": not_temiz,
                "language": {"raw": ", ".join(adlar)} if adlar else None,
                **({"sourceDiff": {"quotaNote": {"kaynakta": None,
                                                 "sebep": "eposta-ayiklandi"}}}
                   if eposta_vardi else {}),
            }
            rec.update(fam_fields(rec["iscedCode"]))
            # The faculty is not a field of the record (no other source has one)
            # but it is real information the student searches by, so it goes
            # into the search text.
            rec["search"] = make_search(rec, extra=fakulte)
            agreements.append(rec)
    if atlanan:
        print(f"  marmara: Erasmus kodundan ülkesi okunamayan {atlanan} satır atlandı",
              file=sys.stderr)
    return agreements


# ═══════════════════════════════ ESOGU ═══════════════════════════════

# Levels arrive as four separate numeric columns instead of one cell. A zero
# means the level is not open; the site needs the number, so it is kept as a
# string rather than flattened to a boolean.
ESOGU_SEVIYE = [(6, "onlisans", "ÖL"), (7, "lisans", "L"),
                (8, "yukseklisans", "YL"), (9, "doktora", "D")]


def parse_esogu(ws):
    """Single sheet, one row per agreement, faculty cells merged per block.

    Columns: A faculty, B department, C code, D university, E country,
    F website, G-J quotas per level, K staff teaching, L staff training.

    The faculty cell is filled only on the first row of each block (16 of 259
    rows), so it is carried forward. The same is true of the department cell.
    Blank rows separate the blocks and are skipped.

    Unlike Marmara this source HAS a country column, so the same cross-check
    as MAKU applies: the Erasmus prefix wins and the correction is recorded.
    """
    agreements = []
    fakulte = bolum = ""
    duzeltme, tamamlama, fakulteden = [], [], 0
    for satir in ws.iter_rows(min_row=2, values_only=True):
        if satir and clean(satir[0]):
            yeni = _bosluk_sadelestir(clean(satir[0]))
            if yeni != fakulte:
                # The carried-over department is dropped at a faculty boundary.
                # Without this, a block whose first row has no department would
                # SILENTLY inherit the last department of the previous faculty
                # and put an agreement under the wrong programme. It has not
                # happened in this file, but nothing in the sheet prevents it.
                bolum = ""
            fakulte = yeni
        if satir and len(satir) > 1 and clean(satir[1]):
            bolum = _bosluk_sadelestir(clean(satir[1]))
        kod = clean(satir[2]) if satir and len(satir) > 2 else ""
        if not kod:
            continue
        university = clean(satir[3])
        if not university or not re.search(r"[A-Za-zĞÜŞİÖÇğüşıöç]", university):
            continue

        country = clean(satir[4])
        onekten = ulke_onekten(kod)
        iz = {}
        if onekten and country and tr_upper(onekten) != tr_upper(country):
            duzeltme.append((university[:44], country, onekten))
            iz["country"] = {"kaynakta": country, "sebep": "erasmus-kodu"}
            country = onekten
        elif not country:
            if not onekten:
                continue
            tamamlama.append((university[:44], onekten))
            iz["country"] = {"kaynakta": None, "sebep": "erasmus-kodu"}
            country = onekten

        seviyeler, kontenjan = {}, []
        for sutun, anahtar, etiket in ESOGU_SEVIYE:
            ham = clean(satir[sutun]) if len(satir) > sutun else ""
            if ham.isdigit() and int(ham) > 0:
                seviyeler[anahtar] = ham
                kontenjan.append(f"{ham}{etiket}")

        # An agreement with no department belongs to the faculty as a whole.
        # Showing the faculty name is the honest reading, but the student must
        # be able to tell it apart from a real department name, so it leaves a
        # trace like every other departure from the source.
        if not bolum and fakulte:
            iz["department"] = {"kaynakta": None, "sebep": "fakulte-adi"}
            fakulteden += 1

        site = clean(satir[5]) if len(satir) > 5 else ""
        rec = {
            "country": tr_upper(country),
            "erasmusCode": kod,
            "validity": None,
            "university": university,
            "iscedCode": "",
            "department": bolum or fakulte,
            "levels": seviyeler,
            # Four columns rendered as one string. This is FORMATTING, not a
            # departure: every number below comes from the sheet unchanged and
            # the notation is the one Marmara's own source uses, so the two
            # universities read alike on the card.
            "quotaStudy": ", ".join(kontenjan),
            "quotaInternship": "",
            "quotaStaffTeach": clean(satir[10]) if len(satir) > 10 else "",
            "quotaStaffTrain": clean(satir[11]) if len(satir) > 11 else "",
            "quotaNote": "",
            "language": None,
            **({"website": site} if site.startswith("http") else {}),
            **({"sourceDiff": iz} if iz else {}),
        }
        rec.update(fam_fields(rec["iscedCode"]))
        rec["search"] = make_search(rec, extra=fakulte)
        agreements.append(rec)
    for ad, yazan, dogru in duzeltme:
        print(f"  esogu: ülke düzeltildi · {ad}: {yazan} -> {dogru} "
              f"(Erasmus kodunun ön eki esas alındı)", file=sys.stderr)
    for ad, dogru in tamamlama:
        print(f"  esogu: ülke boştu, ön ekten dolduruldu · {ad}: {dogru}", file=sys.stderr)
    if fakulteden:
        print(f"  esogu: bölümü yazılmayan {fakulteden} kayıtta fakülte adı gösteriliyor",
              file=sys.stderr)
    return agreements


# ═══════════════════════════ University registry ════════════════════

UNIVERSITIES = [
    {
        "id": "maku",
        "nameTr": "Burdur Mehmet Akif Ersoy Üniversitesi",
        "nameEn": "Burdur Mehmet Akif Ersoy University",
        "abbr": "MAKÜ",
        "monogram": "M",
        "creditTr": "Veri: MAKÜ Uluslararası İlişkiler Koordinatörlüğü",
        "creditEn": "Data: MAKÜ International Relations Office",
        "hasGuide": True,
        "local": ROOT / "lokal" / "maku" / "maku-ka131-ikili-anlasmalar.xlsx",
        # The file name is versioned; when it breaks, the current link comes
        # from listUrl below. listUrl is also SHOWN ON THE SITE: a reader must
        # be able to reach the institution's own list in one click and compare.
        "pullUrl": "https://depo2.mehmetakif.edu.tr/storage/iro/contents/24317/29_24317_2026-03-31-10-48-36-47697_ka-3--ikili-anlasmalar-surum-2-22-26.xlsx",
        "listUrl": "https://iro.mehmetakif.edu.tr/content/24317/1/erasmus-ka131-ikili-anlasmalari",
        "sheet": 0,
        "parser": parse_maku,
    },
    {
        "id": "marmara",
        "nameTr": "Marmara Üniversitesi",
        "nameEn": "Marmara University",
        "abbr": "MÜ",
        "monogram": "MÜ",
        "creditTr": "Veri: Marmara Üniversitesi Uluslararası İlişkiler Koordinatörlüğü",
        "creditEn": "Data: Marmara University International Relations Office",
        "hasGuide": False,
        "local": ROOT / "lokal" / "marmara" / "marmara-ka131-ikili-anlasmalar.xlsx",
        # The file sits behind a share link with an opaque token, so there is no
        # stable direct URL to pull from. The list page is the durable address;
        # a new file is fetched by hand from there.
        "pullUrl": None,
        "listUrl": "https://cloud.marmara.edu.tr/s/1kNVSU18NKjHkdy?openfile=true",
        # None = the parser wants the WHOLE workbook, not one sheet: the
        # agreements are spread over 27 faculty sheets.
        "sheet": None,
        "parser": parse_marmara,
    },
    {
        "id": "esogu",
        "nameTr": "Eskişehir Osmangazi Üniversitesi",
        "nameEn": "Eskisehir Osmangazi University",
        "abbr": "ESOGÜ",
        "monogram": "E",
        "creditTr": "Veri: ESOGÜ Uluslararası İlişkiler Birimi",
        "creditEn": "Data: ESOGU International Relations Office",
        "hasGuide": False,
        "local": ROOT / "lokal" / "esogu" / "esogu-ikili-anlasmalar.xlsx",
        "pullUrl": None,
        "listUrl": "https://iro.ogu.edu.tr/Sayfa/Index/22/ikili-anlasmalar",
        "sheet": 0,
        "parser": parse_esogu,
    },
]


def load_ws(uni, pull):
    """Returns what the parser asked for: one worksheet, or the whole workbook.

    `sheet: None` means the parser needs every sheet (Marmara spreads its
    agreements over 27 faculty sheets). Passing the workbook instead of
    silently handing over the first sheet keeps the choice visible here rather
    than hidden inside each parser.
    """
    src = uni["local"]
    if pull:
        if not uni.get("pullUrl"):
            # Not every source has a stable direct URL. Saying so is better
            # than a generic failure: the file is fetched by hand from listUrl.
            print(f"  {uni['id']}: doğrudan indirme adresi yok, yerel snapshot "
                  f"kullanılıyor · yeni dosya {uni.get('listUrl')} sayfasından alınır",
                  file=sys.stderr)
        else:
            try:
                print(f"  çekiliyor: {uni['pullUrl']}", file=sys.stderr)
                data = urllib.request.urlopen(uni["pullUrl"], timeout=60).read()
                src.write_bytes(data)
            except Exception as e:
                print(f"  UYARI: {uni['id']} canlı çekilemedi ({e}), yerel snapshot "
                      f"kullanılıyor", file=sys.stderr)
    wb = openpyxl.load_workbook(src, data_only=True)
    if uni["sheet"] is None:
        return wb
    return wb[uni["sheet"]] if isinstance(uni["sheet"], str) else wb.worksheets[uni["sheet"]]


def build(pull=False):
    SITE.mkdir(parents=True, exist_ok=True)
    registry = []
    now = datetime.now().isoformat(timespec="seconds")

    for uni in UNIVERSITIES:
        print(f"{uni['id']}: {uni['local'].name}", file=sys.stderr)
        agreements = uni["parser"](load_ws(uni, pull))
        birlesen, beraberlik = unify_department_spelling(agreements)
        if birlesen or beraberlik:
            print(f"  {uni['id']}: bölüm yazımı · {birlesen} kayıt birleştirildi, "
                  f"{beraberlik} grupta beraberlik (dokunulmadı)", file=sys.stderr)
        countries = sorted({a["country"] for a in agreements})
        families = {}
        for a in agreements:
            if a["iscedFamily"]:
                families[a["iscedFamily"]] = families.get(a["iscedFamily"], 0) + 1
        out = {
            "generatedAt": now,
            "source": "live" if pull else "local-snapshot",
            "count": len(agreements),
            "countries": countries,
            "iscedFamilies": [
                {"code": k, "tr": ISCED_FAMILY[k][0], "en": ISCED_FAMILY[k][1], "count": v}
                for k, v in sorted(families.items())
            ],
            "agreements": agreements,
        }
        path = SITE / f"data-{uni['id']}.json"
        path.write_text(json.dumps(out, ensure_ascii=False), encoding="utf-8")

        bil = sum(1 for a in agreements if "bilgisayar" in a["search"] or "computer" in a["search"])
        print(f"  OK kayıt={len(agreements)} ülke={len(countries)} "
              f"aile={len(families)} bilgisayar/computer={bil} -> {path.name} "
              f"({path.stat().st_size // 1024} KB)", file=sys.stderr)

        registry.append({
            "id": uni["id"], "nameTr": uni["nameTr"], "nameEn": uni["nameEn"],
            "abbr": uni["abbr"], "monogram": uni["monogram"],
            # The institution's own list. Shown on the site so a reader can go
            # to the source and compare · this is what makes "we changed
            # things" checkable instead of a claim.
            "sourceUrl": uni.get("listUrl"),
            # How many records are shown differently from the source. Counted,
            # not written: the number moves with the data.
            "diffCount": sum(1 for a in agreements if a.get("sourceDiff")),
            "creditTr": uni["creditTr"], "creditEn": uni["creditEn"],
            "hasGuide": uni["hasGuide"],
            "count": len(agreements), "countries": len(countries),
            "universities": len({a["university"] for a in agreements}),
            "generatedAt": now,
        })

    (SITE / "universities.json").write_text(
        json.dumps(registry, ensure_ascii=False), encoding="utf-8")
    print(f"universities.json: {len(registry)} üniversite", file=sys.stderr)
    write_sitemap(registry)


ALAN_ADI = "https://exchangeatlas.org"


def write_sitemap(registry):
    """Arama motorlarına sayfa listesi · kayıttan üretiliyor.

    Elle yazılsa üniversite eklendiğinde güncellenmesi unutulurdu ve unutulduğu
    gün hiçbir şey kırılmazdı: Site çalışmaya devam eder, yalnız yeni sayfalar
    aranamaz olurdu. Kayıttan üretmek o boşluğu tümüyle kaldırıyor.
    """
    # Adresler UZANTISIZ · ölçüldü, kanonik yayın `.html`'i uzantısıza 308
    # ile yönlendiriyor:
    #
    #   /agreements.html?uni=maku  →  308  →  /agreements?uni=maku
    #
    # Yani sitemap'teki her adres bir yönlendirmeydi ve arama motoruna
    # sayfanın gerçek adresi hiç bildirilmiyordu. Giriş sayfası da öyleydi:
    # `/index.html` → 308 → `/`, üstelik sayfanın kendi `canonical`'ı zaten
    # `/` diyordu · sitemap ile sayfa birbirini yalanlıyordu.
    #
    # Adresler mutlak ve `ALAN_ADI` ile üretiliyor, yani bu liste KANONİK
    # YAYINI tarif ediyor · deponun herhangi bir statik sunucuya kopyalanmış
    # hâlini değil. Uzantısız biçim o yayının gerçek biçimi.
    yollar = ["/"]
    for uni in registry:
        yollar.append(f"/agreements?uni={uni['id']}")
        if uni.get("hasGuide"):
            yollar.append(f"/guide?uni={uni['id']}")

    bugun = datetime.now().date().isoformat()
    # `if y != "/"` KALDIRILDI. Giriş sayfası listeden atılıp yerine
    # `/index.html` konuyordu; o adres de yönlendiriliyordu. Artık giriş
    # sayfası gerçek adresiyle bildiriliyor.
    satirlar = "\n".join(
        f"  <url><loc>{ALAN_ADI}{y.replace('&', '&amp;')}</loc>"
        f"<lastmod>{bugun}</lastmod></url>"
        for y in yollar
    )
    (SITE / "sitemap.xml").write_text(
        '<?xml version="1.0" encoding="UTF-8"?>\n'
        '<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">\n'
        f"{satirlar}\n</urlset>\n", encoding="utf-8")
    print(f"sitemap.xml: {len(yollar)} adres", file=sys.stderr)


def check_guide_links():
    """guide.html'deki dış linklere HEAD at; kırık olanları uyar (build'i durdurmaz).

    Linkler guide.html'den okunur ki betik ile sayfa arasında liste drift'i olmasın.

    Yalnız `<a href>` bakılıyor. `<link rel="canonical">` de bir href taşıyor ama
    kullanıcının tıklayacağı bir bağlantı değil, sayfanın kendi adresini bildiren
    üstveri · üstelik site yayına çıkana kadar 404 veriyor ve her derlemede
    yanlış alarm üretiyordu.
    """
    guide = SITE / "guide.html"
    if not guide.exists():
        return
    urls = sorted(set(re.findall(r'<a[^>]+href="(https?://[^"]+)"',
                                 guide.read_text(encoding="utf-8"))))
    broken = 0
    for url in urls:
        try:
            req = urllib.request.Request(url, method="HEAD",
                                         headers={"User-Agent": "Mozilla/5.0 (atlas-linkcheck)"})
            code = urllib.request.urlopen(req, timeout=15).status
        except urllib.error.HTTPError as e:
            if e.code == 405:
                try:
                    code = urllib.request.urlopen(url, timeout=15).status
                except Exception:
                    code = e.code
            else:
                code = e.code
        except Exception as e:
            print(f"  UYARI erişilemedi: {url} ({e})", file=sys.stderr)
            broken += 1
            continue
        if code >= 400:
            print(f"  KIRIK LINK ({code}): {url}", file=sys.stderr)
            broken += 1
    status = f"{broken} kırık" if broken else "hepsi sağlam"
    print(f"rehber link kontrolü: {len(urls)} dış link, {status}", file=sys.stderr)


if __name__ == "__main__":
    build(pull="--pull" in sys.argv)
    try:
        check_guide_links()
    except Exception as e:  # link kontrolü hiçbir koşulda build'i bozmasın (offline vb.)
        print(f"rehber link kontrolü atlandı: {e}", file=sys.stderr)
